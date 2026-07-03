# Exporter Excel: menulis hasil simulasi ke workbook openpyxl

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Konstanta gaya ----
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")  # RGB(31,78,121)
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True)
TOTAL_FONT = Font(bold=True)
_THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

NUM = "#,##0"        # angka Rupiah (Rp Juta)
PCT = "0.00%"        # persen
RATIO = "0.00"       # rasio (mis. DSCR)


def generate(data, years, capex_result, financing_result, ap_result,
             pnl_result, cf_result, metric_result) -> bytes:
    """Bangun workbook 6 sheet dari hasil simulasi dan kembalikan sebagai bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # buang sheet default

    # Tahun-tahun relevan per konteks
    capex_years = [y for y in years if capex_result["total_per_tahun"].get(y)]
    fin_years = _financing_years(financing_result)
    active_years = _active_years(
        years, capex_result, ap_result, financing_result, pnl_result, cf_result
    )

    _sheet_asumsi(wb.create_sheet("Asumsi"), data)
    _sheet_capex(wb.create_sheet("Capex"), capex_result, capex_years)
    _sheet_pembiayaan(wb.create_sheet("Pembiayaan"), data, financing_result, fin_years)
    _sheet_pnl(wb.create_sheet("P&L"), pnl_result, active_years)
    _sheet_arus_kas(wb.create_sheet("Arus Kas"), cf_result, active_years)
    _sheet_metrics(wb.create_sheet("Metrics"), metric_result)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Helper umum
# ---------------------------------------------------------------------------
def _header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
    return c


def _write_year_row(ws, row, label, values, year_cols, fmt=NUM, bold=False,
                    start_col=2):
    """Tulis label di kolom A dan nilai per tahun mulai start_col."""
    lc = ws.cell(row=row, column=1, value=label)
    if bold:
        lc.font = TOTAL_FONT
    for i, y in enumerate(year_cols):
        c = ws.cell(row=row, column=start_col + i, value=values.get(y))
        c.number_format = fmt
        if bold:
            c.font = TOTAL_FONT
    return row + 1


def _finalize(ws, n_cols):
    ws.column_dimensions["A"].width = 35
    for col in range(2, max(n_cols, 2) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "B2"


def _financing_years(financing_result):
    ys = set()
    for komp_data in financing_result["per_komponen"].values():
        ys.update(komp_data.keys())
    return sorted(ys)


def _active_years(years, capex_result, ap_result, financing_result,
                  pnl_result, cf_result):
    active = set()
    for y in years:
        if capex_result["total_per_tahun"].get(y):
            active.add(y)
        if ap_result["per_tahun"].get(y):
            active.add(y)
        ft = financing_result["total_per_tahun"][y]
        if any(ft[k] for k in ("drawdown", "bunga", "pokok", "anuitas")):
            active.add(y)
        p = pnl_result[y]
        if any(p[k] for k in ("ap_revenue", "opex", "ebitda", "depresiasi",
                              "ebit", "bunga", "ebt", "pajak", "eat")):
            active.add(y)
        c = cf_result[y]
        if any(c[k] for k in ("cfo", "cfi", "cff", "net_cf")):
            active.add(y)
    return sorted(active)


# ---------------------------------------------------------------------------
# SHEET 1 — Asumsi
# ---------------------------------------------------------------------------
def _sheet_asumsi(ws, data):
    ws["A1"] = "SIMULASI KPBU-AP — ASUMSI"
    ws["A1"].font = TITLE_FONT

    row = 3
    _header_cell(ws, row, 2, "Asumsi Umum")
    _header_cell(ws, row, 3, "Nilai")
    row += 1
    g = data.general
    umum = [
        ("Inflasi", g.inflasi, PCT),
        ("Bunga", g.bunga, PCT),
        ("WACC", g.wacc, PCT),
        ("PPN", g.ppn, PCT),
        ("PPh Badan", g.pph_badan, PCT),
        ("Kurs", g.kurs, NUM),
        ("Porsi Ekuitas", g.porsi_ekuitas, PCT),
        ("Masa Pinjaman", g.masa_pinjaman, "0"),
    ]
    for label, val, fmt in umum:
        ws.cell(row=row, column=2, value=label)
        c = ws.cell(row=row, column=3, value=val)
        c.number_format = fmt
        row += 1

    row += 1
    headers = ["Nama", "Tipe", "Biaya (Rp Juta)", "Masa Konstruksi",
               "Mulai Konstruksi", "Masa Operasi"]
    for i, h in enumerate(headers):
        _header_cell(ws, row, 1 + i, h)
    row += 1
    for komp in data.komponen:
        ws.cell(row=row, column=1, value=komp.nama)
        ws.cell(row=row, column=2, value=komp.tipe)
        ws.cell(row=row, column=3, value=komp.biaya_konstruksi_juta).number_format = NUM
        ws.cell(row=row, column=4, value=komp.masa_konstruksi)
        ws.cell(row=row, column=5, value=komp.mulai_konstruksi)
        ws.cell(row=row, column=6, value=komp.masa_operasi)
        row += 1

    row += 1
    _header_cell(ws, row, 2, "Asumsi KPBU")
    _header_cell(ws, row, 3, "Nilai")
    row += 1
    kpbu = [
        ("AP Start Year", data.kpbu.ap_start_year, "0"),
        ("Masa Konsesi AP", data.kpbu.ap_konsesi, "0"),
        ("AP Coverage", data.kpbu.ap_coverage, PCT),
        ("Biaya O&M", data.biaya_om_persen, PCT),
    ]
    for label, val, fmt in kpbu:
        ws.cell(row=row, column=2, value=label)
        c = ws.cell(row=row, column=3, value=val)
        c.number_format = fmt
        row += 1

    _finalize(ws, 6)


# ---------------------------------------------------------------------------
# SHEET 2 — Capex
# ---------------------------------------------------------------------------
def _sheet_capex(ws, capex_result, capex_years):
    _header_cell(ws, 1, 1, "Komponen")
    _header_cell(ws, 1, 2, "Total")
    for i, y in enumerate(capex_years):
        _header_cell(ws, 1, 3 + i, y)

    row = 2
    for nama, per_tahun in capex_result["per_komponen"].items():
        ws.cell(row=row, column=1, value=nama)
        total = sum(v["total"] for v in per_tahun.values())
        ws.cell(row=row, column=2, value=total).number_format = NUM
        for i, y in enumerate(capex_years):
            val = per_tahun.get(y, {}).get("total", 0.0)
            ws.cell(row=row, column=3 + i, value=val).number_format = NUM
        row += 1

    # Baris TOTAL semua komponen
    lc = ws.cell(row=row, column=1, value="TOTAL")
    lc.font = TOTAL_FONT
    tc = ws.cell(row=row, column=2, value=capex_result["grand_total"])
    tc.number_format = NUM
    tc.font = TOTAL_FONT
    for i, y in enumerate(capex_years):
        c = ws.cell(row=row, column=3 + i,
                    value=capex_result["total_per_tahun"].get(y, 0.0))
        c.number_format = NUM
        c.font = TOTAL_FONT

    _finalize(ws, 2 + len(capex_years))


# ---------------------------------------------------------------------------
# SHEET 3 — Pembiayaan
# ---------------------------------------------------------------------------
def _sheet_pembiayaan(ws, data, financing_result, fin_years):
    _header_cell(ws, 1, 1, "Komponen / Pos")
    for i, y in enumerate(fin_years):
        _header_cell(ws, 1, 2 + i, y)

    row = 2
    kategori = [
        ("Drawdown", "drawdown"),
        ("Saldo", "saldo_akhir"),
        ("Bunga", "bunga"),
        ("Pokok", "pokok"),
        ("Anuitas", "anuitas"),
    ]
    for komp in data.komponen:
        per_tahun = financing_result["per_komponen"].get(komp.nama, {})
        lc = ws.cell(row=row, column=1, value=komp.nama)
        lc.font = TOTAL_FONT
        row += 1
        for label, key in kategori:
            values = {y: per_tahun.get(y, {}).get(key, 0.0) for y in fin_years}
            row = _write_year_row(ws, row, "  " + label, values, fin_years)

    # Baris Total per kategori
    total = financing_result["total_per_tahun"]
    saldo_total = {
        y: sum(
            financing_result["per_komponen"][n].get(y, {}).get("saldo_akhir", 0.0)
            for n in financing_result["per_komponen"]
        )
        for y in fin_years
    }
    totals = [
        ("Total Drawdown", {y: total[y]["drawdown"] for y in fin_years}),
        ("Total Saldo", saldo_total),
        ("Total Bunga", {y: total[y]["bunga"] for y in fin_years}),
        ("Total Pokok", {y: total[y]["pokok"] for y in fin_years}),
        ("Total Anuitas", {y: total[y]["anuitas"] for y in fin_years}),
    ]
    for label, values in totals:
        row = _write_year_row(ws, row, label, values, fin_years, bold=True)

    _finalize(ws, 1 + len(fin_years))


# ---------------------------------------------------------------------------
# SHEET 4 — P&L
# ---------------------------------------------------------------------------
def _sheet_pnl(ws, pnl_result, active_years):
    _header_cell(ws, 1, 1, "Pos")
    for i, y in enumerate(active_years):
        _header_cell(ws, 1, 2 + i, y)

    baris = [
        ("AP Revenue", "ap_revenue"),
        ("Opex", "opex"),
        ("EBITDA", "ebitda"),
        ("Depresiasi", "depresiasi"),
        ("EBIT", "ebit"),
        ("Bunga", "bunga"),
        ("EBT", "ebt"),
        ("Pajak", "pajak"),
        ("EAT", "eat"),
    ]
    row = 2
    for label, key in baris:
        values = {y: pnl_result[y][key] for y in active_years}
        bold = label in ("EBITDA", "EBIT", "EBT", "EAT")
        row = _write_year_row(ws, row, label, values, active_years, bold=bold)

    _finalize(ws, 1 + len(active_years))


# ---------------------------------------------------------------------------
# SHEET 5 — Arus Kas
# ---------------------------------------------------------------------------
def _sheet_arus_kas(ws, cf_result, active_years):
    _header_cell(ws, 1, 1, "Pos")
    for i, y in enumerate(active_years):
        _header_cell(ws, 1, 2 + i, y)

    baris = [
        ("CFO", "cfo", NUM),
        ("CFI", "cfi", NUM),
        ("CFF", "cff", NUM),
        ("Net Cash Flow", "net_cf", NUM),
        ("Kas Awal", "kas_awal", NUM),
        ("Kas Akhir", "kas_akhir", NUM),
        ("FCFP", "fcfp", NUM),
        ("FCFE", "fcfe", NUM),
        ("DSCR", "dscr", RATIO),
    ]
    row = 2
    for label, key, fmt in baris:
        values = {y: cf_result[y][key] for y in active_years}
        bold = label in ("Net Cash Flow", "FCFP", "FCFE")
        row = _write_year_row(ws, row, label, values, active_years, fmt=fmt,
                              bold=bold)

    _finalize(ws, 1 + len(active_years))


# ---------------------------------------------------------------------------
# SHEET 6 — Metrics
# ---------------------------------------------------------------------------
def _sheet_metrics(ws, metric_result):
    ws["A1"] = "RINGKASAN METRICS"
    ws["A1"].font = TITLE_FONT

    summary = [
        ("IRR Proyek", metric_result["irr_proyek"], PCT),
        ("IRR Ekuitas", metric_result["irr_ekuitas"], PCT),
        ("NPV (Rp Juta)", metric_result["npv_proyek"], NUM),
        ("Payback Proyek (tahun)", metric_result["payback_proyek_tahun"], "0"),
        ("Payback Ekuitas (tahun)", metric_result["payback_ekuitas_tahun"], "0"),
        ("DSCR Min", metric_result["dscr_min"], RATIO),
        ("DSCR Avg", metric_result["dscr_avg"], RATIO),
    ]
    row = 2
    for label, val, fmt in summary:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = TOTAL_FONT
        c = ws.cell(row=row, column=3, value=val)
        c.number_format = fmt
        row += 1

    # Tabel kumulatif FCFP & FCFE per tahun
    row = 11
    _header_cell(ws, row, 1, "Tahun")
    _header_cell(ws, row, 2, "FCFP Kumulatif")
    _header_cell(ws, row, 3, "FCFE Kumulatif")
    row += 1
    fcfp_kum = metric_result["fcfp_kumulatif"]
    fcfe_kum = metric_result["fcfe_kumulatif"]
    for y in sorted(fcfp_kum.keys()):
        ws.cell(row=row, column=1, value=y)
        ws.cell(row=row, column=2, value=fcfp_kum[y]).number_format = NUM
        ws.cell(row=row, column=3, value=fcfe_kum[y]).number_format = NUM
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "B2"
